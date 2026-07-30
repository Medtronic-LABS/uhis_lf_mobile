#!/usr/bin/env python3
"""Merge extract_translations.py's output with the best-effort Bangla drafts
in manual_bangla.py, then write:
  - assets/translations/strings.json   (app runtime asset)
  - store_assets/translations_for_brac.xlsx   (BRAC review handoff)

Usage:
    .venv-tools/bin/python3 scripts/build_translations.py
"""
import json
import subprocess
import sys

sys.path.insert(0, "scripts")
from manual_bangla import MANUAL_BANGLA  # noqa: E402

JSON_OUT = "assets/translations/strings.json"
XLSX_OUT = "store_assets/translations_for_brac.xlsx"


def main():
    raw = subprocess.run(
        [sys.executable, "scripts/extract_translations.py"],
        capture_output=True,
        text=True,
        check=True,
    )
    report = json.loads(raw.stdout)
    entries = report["entries"]

    missing_after_merge = []
    for e in entries:
        if not e.get("bangla"):
            bangla = MANUAL_BANGLA.get(e["english"])
            if bangla:
                e["bangla"] = bangla
            else:
                missing_after_merge.append(e["key"])

    if missing_after_merge:
        print(
            f"✗ {len(missing_after_merge)} entries still missing Bangla after merge:",
            file=sys.stderr,
        )
        for k in missing_after_merge[:20]:
            print(f"  {k}", file=sys.stderr)
        sys.exit(1)

    write_json(entries)
    write_xlsx(entries)

    print(f"✓ {len(entries)} entries written to {JSON_OUT}")
    print(f"✓ {len(entries)} rows written to {XLSX_OUT}")
    print(f"  ({report['counts']['unhandled']} declarations left unhandled -- see extract_translations.py output)")


def write_json(entries):
    import os

    os.makedirs("assets/translations", exist_ok=True)
    data = {e["key"]: {"en": e["english"], "bn": e["bangla"]} for e in entries}
    with open(JSON_OUT, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=1, sort_keys=True)
        f.write("\n")


def write_xlsx(entries):
    import os

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.protection import SheetProtection

    os.makedirs("store_assets", exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Translations"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B2B5E", end_color="1B2B5E", fill_type="solid")
    ws.append(["code", "English", "Bangla"])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for e in sorted(entries, key=lambda e: e["key"]):
        ws.append([e["key"], e["english"], e["bangla"]])

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 60
    ws.freeze_panes = "A2"

    # Lock the code column against edits: sheet protection is on, and every
    # cell is locked by default -- we only need to explicitly UNLOCK the
    # English/Bangla columns so BRAC can edit those while "code" stays
    # read-only.
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=3):
        for cell in row:
            cell.protection = cell.protection.copy(locked=False)
    ws.protection = SheetProtection(sheet=True, password=None)
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.sort = False
    ws.protection.autoFilter = False

    wb.save(XLSX_OUT)


if __name__ == "__main__":
    main()
