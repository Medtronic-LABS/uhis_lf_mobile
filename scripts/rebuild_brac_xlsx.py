#!/usr/bin/env python3
"""Regenerate store_assets/translations_for_brac.xlsx from the already-built
assets/translations/strings.json, without re-running extract_translations.py
(which can no longer safely re-parse app_strings.dart once its declarations
have been rewritten to call getTranslatedString -- see rewrite_app_strings.py).

Usage:
    .venv-tools/bin/python3 scripts/rebuild_brac_xlsx.py
"""
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.protection import SheetProtection

JSON_IN = "assets/translations/strings.json"
XLSX_OUT = "store_assets/translations_for_brac.xlsx"


def main():
    with open(JSON_IN, encoding="utf-8") as f:
        data = json.load(f)

    os.makedirs("store_assets", exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Translations"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B2B5E", end_color="1B2B5E", fill_type="solid")
    ws.append(["code", "English", "Bangla"])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for key in sorted(data):
        ws.append([key, data[key]["en"], data[key]["bn"]])

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 60
    ws.freeze_panes = "A2"

    for row in ws.iter_rows(min_row=2, min_col=2, max_col=3):
        for cell in row:
            cell.protection = cell.protection.copy(locked=False)
    ws.protection = SheetProtection(sheet=True, password=None)
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.sort = False
    ws.protection.autoFilter = False

    wb.save(XLSX_OUT)
    print(f"✓ {len(data)} rows written to {XLSX_OUT}")


if __name__ == "__main__":
    main()
